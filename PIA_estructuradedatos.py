
import datetime
import sys
import json
import csv
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import sqlite3
from sqlite3  import Error
import os
import warnings

warnings.filterwarnings("ignore", category=DeprecationWarning)

FORMATO_FECHA = "%m/%d/%Y"
fecha_actual = datetime.datetime.now().date()



def inicializar_db(conn):
    try: 
        cursor = conn.cursor()
        with conn:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Cliente (
                    id_cliente INTEGER PRIMARY KEY,
                    nombre_cliente TEXT NOT NULL,
                    apellido TEXT NOT NULL
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Sala (
                    id_sala INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre_sala TEXT NOT NULL,
                    cupo INTEGER NOT NULL
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Turno (
                    id_turno INTEGER NOT NULL PRIMARY KEY,
                    turno TEXT NOT NULL
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Reservacion (
                    folio INTEGER PRIMARY KEY,
                    id_cliente INTEGER,
                    id_sala INTEGER,
                    fecha DATE NOT NULL,
                    turno INTEGER NOT NULL,
                    evento INTEGER NOT NULL,
                    cancelado INTEGER NOT NULL,
                    FOREIGN KEY (id_turno) REFERENCES Turno (id_turno),
                    FOREIGN KEY (id_cliente) REFERENCES Cliente (id_cliente),
                    FOREIGN KEY (id_sala) REFERENCES Sala (id_sala)
                )
            ''')
            turnos = ("Mañana", "Tarde", "Noche")
            for nombre_turno in turnos:
                cursor.execute(''' INSERT INTO Turno (id_turno, turno)
                                VALUES (?, ?)''', (turnos.index(nombre_turno) + 1, nombre_turno))


    except sqlite3.Error as e:
        print(e)

def conectar_db():
    while True:
        if os.path.exists("archivo.db"):
            print(f" Se encontró una versión anterior del archivo: archivo.db. \n")
            respuesta = input("¿Desea recuperar este estado anterior y continuar? (s/n):").strip().upper()
            
            if respuesta.isalpha() and respuesta == 'S':
                try:
                    conn = sqlite3.connect("archivo.db")
                    print("\nConexión exitosa a la base de datos anterior.\n")
                    return conn
                
                except sqlite3.Error as e:
                    print(f" Error al intentar conectar con la base de datos existente: {e}")
                    print("\nSe iniciará un estado vacío por seguridad.")
                    os.remove("archivo.db") 
                    conn = sqlite3.connect("archivo.db")
                    return inicializar_db(conn)

            elif respuesta.isalpha() and respuesta == 'N':
                
                print("No se recuperará el estado anterior.\n")
                confirmar = input("¿Desea iniciar con un estado inicial vacío (esto SOBREESCRIBIRÁ la versión anterior si guarda)? (s/n): ").strip().upper()
                
                if confirmar.isalpha() and confirmar == 'S':
                    print("\nEliminando el archivo anterior para crear uno nuevo.\n\n")
                    try:
                        os.remove("archivo.db")
                    except Exception:
                        print(f"\nOcurrió un problema: {sys.exc_info()[0:2]}\n\n")
                        continue
                    
                    conn = sqlite3.connect("archivo.db")
                    return inicializar_db(conn)
                elif confirmar.isalpha() and confirmar == 'N':
                    print("Programa cancelado.")
                    sys.exit()
                else:
                    print("Esa opcion no existe")
                    continue
            else:
                print("Esa opcion no existe")
                continue         
        else:
            print(f" No se encontró una versión anterior archivo.db .")
            print("Iniciando con un estado inicial vacío.")
    
            conn = sqlite3.connect("archivo.db")
            return inicializar_db(conn)


def generar_folio():
    try: 
        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(folio) FROM Reservacion")
            result = cursor.fetchone()
            
            return (result[0] or 1000) + 1  
    except sqlite3.Error as e:
        print(e)


def mostrar_clientes():
    try:
        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT c.id_cliente, c.apellido, c.nombre_cliente FROM Cliente c ORDER BY c.apellido, c.nombre_cliente")
            cursor_turno = cursor.fetchall()

            if not cursor_turno:
                print("\nNo hay clientes registrados.\n")
                return False
            print(f"{'CLAVE':<10}{'APELLIDO':<15}{'NOMBRE':<15}")
            print("-" * 40)
            for id_cliente, apellido, nombre in  cursor_turno:  
                print(f"{id_cliente:<10}{apellido:<15}{nombre:<15}")
            return True
    except Exception as e:
        print(f"Error al mostrar clientes: {e}")
        return False


def mostrar_salas_disponibles(fecha, turno):
    try:
        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT s.id_sala, s.nombre_sala, s.cupo 
                FROM Sala s
                WHERE s.id_sala NOT IN (
                    SELECT r.id_sala 
                    FROM Reservacion r
                    WHERE r.fecha = ? AND r.turno = ? AND r.cancelado = 0
                )
            ''', (fecha, turno))
            
        salas_disponibles = cursor.fetchall()
        disponibles = []
        
        print(f"{'CLAVE':<10}{'NOMBRE':<16}{'CUPO':<5}")
        print("-" * 40)
        for id_sala, nombre, cupo in salas_disponibles:
            print(f"{id_sala:<10}{nombre:<16}{cupo:<5}")
            disponibles.append(id_sala)
        return disponibles
    except Exception as e:
        print(f"Error al mostrar salas disponibles: {e}")
        return []


def export_csv(fecha):
    datos_a_leer = {}
    with sqlite3.connect("archivo.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
        cursor = conn.cursor()
        cursor.execute("Select r.folio, r.fecha, r.id_sala, r.turno, r.evento, r.id_cliente From Reservacion r WHERE r.fecha = (?) AND r.cancelado = 0", (fecha,))
        reservaciones = cursor.fetchall()
                               

    with open("reservaciones.csv", "w", encoding='utf-8', newline="") as archivo:
        grabador = csv.writer(archivo)
        grabador.writerow(("FOLIO", "FECHA", "SALA", "TURNO", "EVENTO", "CLIENTE",))

        for folio, fecha_db, sala, turno, evento, cliente in reservaciones:
                grabador.writerow((folio, fecha_db, sala, turno, evento, cliente ))
    
    print(f"\nDatos exportados exitosamente a 'reservaciones.csv'")

    with open("reservaciones.csv","r", encoding='utf-8', newline="") as archivo:
        lector = csv.reader(archivo)
        next(lector)

        for folio, fecha, sala, turno, evento, cliente in lector:
                datos_a_leer[int(folio)] = [fecha, sala, turno, evento, cliente]


def export_json(fecha):
    with sqlite3.connect("archivo.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
        cursor = conn.cursor()
        cursor.execute("Select r.folio, r.id_cliente, r.id_sala, r.fecha, r.turno, r.evento From Reservacion r WHERE r.fecha = (?) AND r.cancelado = 0", (fecha,))

        reservaciones = cursor.fetchall()
        
    datos_leer = {}
    for folio, cliente, sala, fecha_db, turno, evento in reservaciones:
        
        datos_leer[folio] = {
        "FOLIO": folio,
        "FECHA": fecha_db.strftime(FORMATO_FECHA),
        "SALA": sala,
        "TURNO": turno,
        "EVENTO": evento,
        "CLIENTE": cliente
            }
        
    with open("reservaciones.json", "w", encoding='utf-8') as archivo:
     json.dump(datos_leer, archivo, indent=4, ensure_ascii= False)

    print(f"\nDatos exportados exitosamente a 'reservaciones.json'")
    print(f"Total de reservaciones exportadas: {len(datos_leer)}")


def export_excel(fecha):
    try:
        with sqlite3.connect("archivo.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            cursor = conn.cursor()
            cursor.execute("Select r.folio, r.id_sala, r.fecha, r.turno, r.evento, r.id_cliente From Reservacion r WHERE r.fecha = (?) AND r.cancelado = 0", (fecha,))

            reservaciones_db = cursor.fetchall()    

            datos_a_leer = {}
    
            for folio, id_sala, fecha_db, turno, evento, id_cliente in reservaciones_db:
                datos_a_leer[folio] = {
                    "FOLIO": folio,
                    "SALA": id_sala,
                    "FECHA": fecha_db,
                    "TURNO": turno,
                    "EVENTO": evento,
                    "CLIENTE": id_cliente
                }
                
            datos_reservaciones = openpyxl.Workbook()
            hoja = datos_reservaciones["Sheet"]
            hoja.title = "Reservaciones"

            Encabezados = ["Folio", "Sala", "Fecha", "Turno", "Evento", "Cliente"]
            for columna, nombre in enumerate(Encabezados, start=1):
                celda = hoja.cell(row=1, column=columna)
                celda.value = nombre
                celda.font = Font(bold=True)
                celda.border = Border(bottom=Side(style='thick', color="000000"))

            centrado = Alignment(horizontal="center", vertical="center")
            fila_actual = 2
                    
            for folio, sala, fecha_db, turno, evento, id_cliente in reservaciones_db:
                hoja.cell(row=fila_actual, column=1, value=folio).alignment = centrado
                hoja.cell(row= fila_actual, column=2, value=sala).alignment = centrado
                hoja.cell(row= fila_actual, column=3, value=fecha_db.strftime(FORMATO_FECHA)).alignment = centrado
                hoja.cell(row= fila_actual, column=4, value=turno).alignment = centrado
                hoja.cell(row= fila_actual, column=5, value=evento).alignment = centrado
                hoja.cell(row= fila_actual, column=6, value= id_cliente).alignment = centrado



                fila_actual += 1

            hoja.column_dimensions['A'].width = 12
            hoja.column_dimensions['B'].width = 20
            hoja.column_dimensions['C'].width = 15
            hoja.column_dimensions['D'].width = 12
            hoja.column_dimensions['E'].width = 25
            hoja.column_dimensions['F'].width = 25
            

            datos_reservaciones.save("Reservaciones.xlsx")
            print("Archivo Excel guardado con exito")
            
    except Exception:
        print(f"\nError al guardar excel: {sys.exc_info()[0:2]}\n")


def menu_exportacion(fecha):
    while True:
            exportar = input("\nDesea exportar esta consulta a otro tipo de archivo? (S/N): ").strip()
            if not exportar.isalpha():
                print("Respuesta no valida")
                continue
            if not (exportar.upper() == "N" or  exportar.upper() == "S"):
                print(f"La opcion {exportar} no esta dentro de las opciones validas. Intenta de nuevo.")
                continue
            if exportar.upper() == "N":
                break
            else:
                print("\n--- MENÚ ---")
                print("[1] CSV")
                print("[2] JSON")
                print("[3] EXCEL")
                print("[4] CANCELAR")

                try:
                    op = input("\nElige una opción: ").strip()
                    if op == "1":   
                        print("\n")
                        export_csv(fecha)
                    elif op == "2":
                        print("\n")
                        export_json(fecha)
                    elif op == "3":
                        print("\n")
                        export_excel(fecha)
                    elif op == "4":
                        print("\n")
                        print("Cancelando Operacion...")
                        break
                    else:
                        print("Opción no válida.")
                except Exception:
                    print(f"Ocurrió un problema: {sys.exc_info()[0:2]}")
                    continue


def reservacion():
    try:
        if not mostrar_clientes():
            return
        while True:
            clave_cliente = input("\nDime tu clave de cliente (enter para salir): ").strip()

            if clave_cliente == "":
                print("Saliendo...\n\n")
                return
            
            if not clave_cliente.isdigit():
                print("\nClave inválida. Intenta de nuevo.\n\n")
                mostrar_clientes()
                continue
            
            clave_cliente = int(clave_cliente)

            with sqlite3.connect("archivo.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT c.id_cliente FROM Cliente c WHERE c.id_cliente = (?)", (clave_cliente,))
                cursor_turno = cursor.fetchall()
                
            if not cursor_turno:
                print("\nClave no registrada. Intenta de nuevo\n\n")
                mostrar_clientes()
                continue
            break

        while True:
            fecha_str = input("\nQue fecha le gustaria reservar (mm/dd/aaaa): ").strip()

            try:
                fecha = datetime.datetime.strptime(fecha_str, FORMATO_FECHA).date()
                if (fecha - fecha_actual).days < 2:
                    print("\nLa fecha debe ser al menos dos días después de hoy.\n")
                    continue

                if fecha.weekday() == 6:  

                    print("\nNo se pueden realizar reservaciones los domingos.\n")
                    posponer_fecha = input("\nDesea intentar con Lunes? (S/N): ").strip().upper()

                    if posponer_fecha.isalpha() and posponer_fecha == "S":
                        fecha += datetime.timedelta(days=(7 - fecha.weekday()))
                        print(f"\nNueva fecha seleccionada: {fecha.strftime(FORMATO_FECHA)}\n")
                        break

                    elif posponer_fecha.isalpha() and posponer_fecha == "N":
                        continue

                    else:
                        print("Esa opcion no existe")
                break
            except ValueError:
                print("\nFormato de fecha incorrecto.\n\n")
                continue
            except Exception:
                print(f"\nOcurrió un problema: {sys.exc_info()[0:2]}\n\n")
                continue

 
        while True:
            print("\nTurnos disponibles: 1) Mañana 2) Tarde 3) Noche\n")

            turno = input("\nSelecciona turno (1-3):  ").strip()
            print("\n\n")
            with sqlite3.connect("archivo.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT turno FROM Turno t WHERE t.id_turno = (?)", (turno))
                cursor_turno = cursor.fetchall()
                
            if not cursor_turno:
                print("\nTurno no encontrado. Intenta de nuevo\n\n")
                continue

            break
        
        disponibles = mostrar_salas_disponibles(fecha, turno)
        if not disponibles:
            print("\nNo hay salas disponibles para ese turno y fecha.\n")
            return
        
        while True:
            clave_sala = input("\nClave de sala a reservar: ").strip()
            if not clave_sala.isdigit() or int(clave_sala) not in disponibles:
                print("\nClave de sala inválida.\n")
                continue
            
            clave_sala = int(clave_sala)
            break

        while True:
            evento = input("Nombre del evento: ").strip()
            if not evento:
                print("\nEl nombre del evento no puede estar vacío.\n")
                continue
            break
        
        folio = generar_folio()

        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO Reservacion (folio, id_cliente, id_sala, fecha, turno, evento, cancelado)
                VALUES (?, ?, ?, ?, ?, ?, 0)
            ''', (folio, clave_cliente, clave_sala, fecha, turno, evento))

        print(f"\nReservación registrada con folio: {folio}\n")
    except ValueError:
        print("\nHubo un error al intentar ingresar un valor incorrecto\n")
        return
    except Exception:
        print(f"\nError en reservación: {sys.exc_info()[0:2]}\n ")


def editar_reservacion():
    try:
        print("Editar nombre de evento de una reservación.")
        fecha_inicial_str = input("Fecha inicial (mm/dd/aaaa): ").strip()
        fecha_final_str = input("Fecha final (mm/dd/aaaa): ").strip()

    
        fecha_inicial = datetime.datetime.strptime(fecha_inicial_str, FORMATO_FECHA).date()
        fecha_final = datetime.datetime.strptime(fecha_final_str, FORMATO_FECHA).date()

        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT r.folio, r.fecha, r.evento FROM " \
            " Reservacion r WHERE r.fecha BETWEEN ? AND ? AND r.cancelado = 0 ", (fecha_inicial, fecha_final))
            eventos = cursor.fetchall()
            
            if not eventos:
                print("No hay eventos en ese rango.")
                return
            
            print(f"{'FOLIO':<10}{'FECHA':<15}{'EVENTO':<15}")
            print("-" * 40)
            for folio, fecha, evento in eventos:
                print(f"{folio:<10}{fecha:<15}{evento:<15}")

            while True:

                folio_seleccionado = input("Folio a editar (enter para cancelar): ").strip()
                if folio_seleccionado == "":
                    print("Cancelado.")
                    return
                if not folio_seleccionado.isdigit():
                    print("Folio inválido.")
                    continue
                folio_seleccionado = int(folio_seleccionado)

                folios_permitidos = []
                for evento in eventos:
                    folio = evento[0]
                    folios_permitidos.append(folio)

                if folio_seleccionado not in folios_permitidos:
                    print("Folio no encontrado")
                break
            
            while True:
                nuevo_nombre = input("Nuevo nombre del evento: ").strip()
                if not nuevo_nombre:
                    print("El nombre del evento no puede estar vacío.")
                    continue

                with sqlite3.connect("archivo.db") as conn:
                    cursor = conn.cursor()
                    cursor.execute("UPDATE Reservacion SET evento=? WHERE folio=?", (nuevo_nombre, folio_seleccionado))
                print("Evento actualizado.")
                break

    except Exception:
        print(f"Error al editar reservación: {sys.exc_info()[0:2]} ")
    except ValueError:
        print("Formato de fecha incorrecto.")
        return


def consultar_reservaciones():
    try:
        fecha_str = input("Fecha a consultar (mm/dd/aaaa): ").strip()
        
        if fecha_str == "":
            fecha = fecha_actual
        else:
            fecha = datetime.datetime.strptime(fecha_str, FORMATO_FECHA).date()
        
        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            cursor.execute ('''SELECT r.folio, s.nombre_sala, r.turno, r.evento, 
                            c.nombre_cliente, c.apellido
                            FROM Reservacion r
                            INNER JOIN Cliente c ON r.id_cliente = c.id_cliente 
                            INNER JOIN Sala s ON r.id_sala = s.id_sala
                            WHERE r.fecha = ? AND r.cancelado = 0 ''', (fecha,))
            reservaciones = cursor.fetchall()
        
        if not reservaciones:
            print(f"No hay reservaciones para la fecha {fecha.strftime(FORMATO_FECHA)}")
            return
            
        print("*" * 80)
        print(f"{'Reservaciones':^76}")
        print("*" * 80)
        print(f"{'Folio':<8}{'Sala':<16}{'Turno':<16}{'Evento':<20}{'Cliente':<24}")
        print("=" * 80)
        for reserva in reservaciones:
            folio, sala, turno, evento, nombre, apellido = reserva
            cliente_nombre = f"{nombre} {apellido}"
            print(f"{folio:<8}{sala:<16}{turno:<16}{evento:<20}{cliente_nombre:<24}")
    
    except sqlite3.Error as e:
            print(e)    
    except ValueError:
            print("Formato de fecha incorrecto.")
            return    
    except Exception:
        print(f"Error al consultar reservaciones: {sys.exc_info()[0:2]} ")
    else:
        menu_exportacion(fecha)



def new_cliente():
    try:
        nombre = input("Nombre: ").strip()
        if not (nombre.replace(" ", "").isalpha()):
            print("Error al agregar cliente, solo se permiten letras. Introduzca nuevamente.")
            return
        apellidos = input("Apellidos: ").strip()
        if not (apellidos.replace(" ", "").isalpha()):
            print("Error al agregar cliente, solo se permiten letras. Introduzca nuevamente.")
            return
        
        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(id_cliente) FROM Cliente")
            result = cursor.fetchone()
            cliente_id = (result[0] or 100) + 1  


        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            
            datos_clientes = (cliente_id, nombre, apellidos)

            cursor.execute(" INSERT INTO Cliente (id_cliente,  nombre_cliente, apellido) VALUES (?, ?, ?) ", datos_clientes)
    except sqlite3.Error as e:
        print(e) 
    except Exception:
        print(f"\nError al registrar cliente: {sys.exc_info()[0:2]}\n")
    else:
        print(f"\nCliente registrado con clave: {cliente_id}\n\n")


def new_sala():
    try:
        nombre = input("Nombre de la sala: ").strip()
        if not (nombre.replace(" ", "").isalpha()) :
            print("\nEl nombre de la sala debe contener solo letras y no puede estar vacio.\n")
            return
        
        cupo = input("Capacidad de la sala: ").strip()
        if not cupo.isdigit() or int(cupo) <= 0:
            print("\nEl cupo debe ser un número entero positivo.\n")
            return 
        
        datos_sala = (nombre, int(cupo))

        with sqlite3.connect("archivo.db") as conn:
            cursor = conn.cursor()
            cursor.execute(" INSERT INTO Sala (nombre_sala, cupo) VALUES (?, ?) ", (datos_sala))
            clave  = cursor.lastrowid
            
    except sqlite3.Error as e:
        print(e)    
    except ValueError:
        print("\nEl cupo debe ser un numero entero positivo\n")
        return
    except Exception:
        print(f"\nError al registrar sala: {sys.exc_info()[0:2]}\n")
    else:
        print(f"\nSala registrada con clave: {clave}\n\n")


def cancelar_rsv():
    fecha_inicial_str = input("Fecha inicial (mm/dd/aaaa): ").strip()
    fecha_final_str = input("Fecha final (mm/dd/aaaa): ").strip()
    try:
        fecha_inicial = datetime.datetime.strptime(fecha_inicial_str, FORMATO_FECHA).date()
        fecha_final = datetime.datetime.strptime(fecha_final_str, FORMATO_FECHA).date()
    except ValueError:
        print("Error: Formato de fecha incorrecto. \n")
        return

    with sqlite3.connect("archivo.db", 
                         detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            cursor = conn.cursor()
            cursor.execute('''SELECT r.folio, r.evento, r.fecha
                           FROM Reservacion r
                           WHERE r.fecha BETWEEN ? AND ? AND r.cancelado = 0 ''', (fecha_inicial, fecha_final))
            eventos = cursor.fetchall()

    if not eventos:
        print("No hay eventos en ese rango.\n")
        return

    print(f"{'FOLIO':<10}{'NOMBRE':<20}{'FECHA'}")
    print("-" * 40)
    for folio, evento ,fecha in eventos:
        print(f"{folio:<10}{evento:<20}{fecha}")

    while True:
        try:
            folio_cancelar = int(input("Ingrese el folio de la reservación a cancelar: "))

            reserva_encontrada = None
            for folio, evento ,fecha in eventos:
                if folio == folio_cancelar:
                    reserva_encontrada = (folio, evento ,fecha)
                    break

            if not reserva_encontrada:
                print("Folio no encontrado.")
                continue

            if (reserva_encontrada[2] - fecha_actual).days < 2:
                print("\nNo se puede cancelar. La reservación debe tener al menos 2 días de anticipación.\n")
                continue

            confirmacion = input("Esta seguro que desea cancelar? (S/N): ").strip().upper()
            if not confirmacion.isalpha():
                    print("respuesta invalida")
                    continue

            if confirmacion== "S":
                with sqlite3.connect("archivo.db") as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                                      UPDATE Reservacion 
                                      SET cancelado = 1
                                      WHERE folio = ?
                                   ''', (folio_cancelar,))
                    print("Reservación cancelada exitosamente.")
                break

            elif confirmacion == "N":
                print("Operacion cancelada")
            break

        except sqlite3.Error as e:
            print(e)   
        except ValueError:
            print("Error: Folio debe ser un número.")


def main():
    conectar_db()
    while True:
        print("\n--- MENÚ ---")
        print("[1] Registrar reservación")
        print("[2] Editar reservación")
        print("[3] Consultar reservaciones")
        print("[4] Registrar cliente")
        print("[5] Registrar sala")
        print("[6] Cancelar una reservación")
        print("[7] Salir")
        try:
            op = input("\nElige una opción: ").strip()
            if op == "1":
                print("\n")
                reservacion()

            elif op == "2":
                print("\n")
                editar_reservacion()

            elif op == "3":
                print("\n")
                consultar_reservaciones()

            elif op == "4":
                print("\n")
                new_cliente()

            elif op == "5":
                print("\n")
                new_sala()

            elif op == "6":
                print("\n")
                cancelar_rsv()

            elif op == "7":
                print("\n")

                try:
                    confirmacion = input("Esta seguro que desea salir? (S/N): ").strip()
                    if not confirmacion.isalpha():
                        continue
                    if not (confirmacion.upper() == "N" or  confirmacion.upper() == "S"):
                        print(f"La opcion {confirmacion} no esta dentro de las opciones validas. Intenta de nuevo.")
                        continue
                    if confirmacion.upper()== "N":
                        continue
                    else:
                        print("Saliendo del programa...")
                        break
                except Exception:
                    print(f"Ocurrió un problema: {sys.exc_info()[0:2]}")
                    continue
            else:
                print("Opción no válida.")

        except ValueError:
            print("La opcion proporcionada no es valida, intente de nuevo...")
            continue
        except Exception:
            print(f"Ocurrió un problema: {sys.exc_info()[0:2]}")
            continue


if __name__ == "__main__":
    main()


