import datetime
import sys
import json
import csv
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment


FORMATO_FECHA = "%d/%m/%Y"
fecha_actual = datetime.datetime.now().date()

#clientes = {}  # clave: [apellidos, nombre]
#salas = {}     # clave: [nombre, cupo]
#reservas = {}  # folio: {cliente, sala, fecha, turno, evento}
#salas_ocupadas= {}  # (sala, fecha, turno): folio

def generar_folio():
    return max(reservas.keys(), default=1000) + 1

def mostrar_clientes():
    if not clientes:
        print("\nTodavía no hay clientes registrados.\n\n")
        return False
    else:
        print("CLAVE\t\tAPELLIDO\tNOMBRE")
        print("-" * 40)
        for clave, datos in sorted(clientes.items(), key=lambda item: (item[1][0], item[1][1])):
            print(f"{clave:<16}{datos[0]:<16}{datos[1]}")
        return True

def mostrar_salas_disponibles(fecha, turno):
    disponibles = []
    print("CLAVE\t\tNOMBRE\t\tCUPO")
    print("-" * 40)
    for clave, (nombre, cupo) in salas.items():
        if (clave, fecha, turno) not in salas_ocupadas:
            print(f"{clave:<16}{nombre:<16}{cupo}")
            disponibles.append(clave)
    return disponibles

def export_csv(fecha):
    datos_a_leer = {}
    with open("reservaciones.csv", "w", encoding='utf-8', newline="") as archivo:
        grabador = csv.writer(archivo)
        grabador.writerow(("FOLIO", "FECHA", "SALA", "TURNO", "EVENTO", "CLIENTE"))

        for folio, datos in reservas.items():
            if datos["fecha"] == fecha:
                sala = salas[datos["sala"]][0]
                cliente_datos = clientes[datos["cliente"]]
                cliente_nombre = (f"{cliente_datos[1]} {cliente_datos[0]}")

                grabador.writerow((
                    folio, 
                    fecha.strftime(FORMATO_FECHA), 
                    sala, 
                    datos['turno'], 
                    datos['evento'], 
                    cliente_nombre
                ))
               
    print(f"\nDatos exportados exitosamente a 'reservaciones.csv'")


    with open("reservaciones.csv","r", encoding='utf-8', newline="") as archivo:
        lector = csv.reader(archivo)
        next(lector)

        for folio, fecha, sala, turno, evento, cliente in lector:
                datos_a_leer[int(folio)] = [fecha, sala, turno, evento, cliente]

def export_json(fecha):
    datos_leer = {}

    for folio, datos in reservas.items():
        if datos["fecha"] == fecha:
           sala = salas[datos["sala"]][0]
           cliente_datos = clientes[datos["cliente"]]
           cliente_nombre = f"{cliente_datos[1]} {cliente_datos[0]}" 

           datos_leer[folio] = {
            "FOLIO": folio,
            "FECHA": fecha.strftime(FORMATO_FECHA),
            "SALA": sala,
            "TURNO": datos['turno'],
            "EVENTO": datos['evento'],
            "CLIENTE": cliente_nombre
              }
        


    with open("reservaciones.json", "w", encoding='utf-8') as archivo:
     json.dump(datos_leer, archivo, indent=4, ensure_ascii= False)

    print(f"\nDatos exportados exitosamente a 'reservaciones.json'")
    print(f"Total de reservaciones exportadas: {len(datos_leer)}")

def export_excel(fecha):
    eventos = {folio: datos for folio, datos in reservas.items() if datos["fecha"] == fecha}
    
    if not eventos:
        print("No hay reservaciones para esa fecha.")
        return
    
    try:
        Reservaciones = openpyxl.Workbook()
        hoja = Reservaciones["Sheet"]
        hoja.title = "Reservaciones"

        Encabezados = ["Folio", "Sala", "Fecha", "Turno", "Evento", "Cliente"]
        for columna, nombre in enumerate(Encabezados, start=1):
            celda = hoja.cell(row=1, column=columna)
            celda.value = nombre
            celda.font = Font(bold=True)
            celda.border = Border(bottom=Side(style='thick', color="000000"))

        centrado = Alignment(horizontal="center", vertical="center")
        fila_actual = 2
        
        for folio, datos in eventos.items():
            sala = salas[datos["sala"]][0]
            cliente_datos = clientes[datos["cliente"]]
            cliente_nombre = f"{cliente_datos[0]}, {cliente_datos[1]}"
           
            valores = [
                folio, 
                sala, 
                datos["fecha"].strftime(FORMATO_FECHA), 
                datos["turno"], 
                datos["evento"], 
                cliente_nombre
            ]
            
            for columna, valor in enumerate(valores, start=1):
                celda = hoja.cell(row=fila_actual, column=columna)
                celda.value = valor
                celda.alignment = centrado
            
            fila_actual += 1

        hoja.column_dimensions['A'].width = 12
        hoja.column_dimensions['B'].width = 20
        hoja.column_dimensions['C'].width = 15
        hoja.column_dimensions['D'].width = 12
        hoja.column_dimensions['E'].width = 25
        hoja.column_dimensions['F'].width = 25

        Reservaciones.save("Reservaciones.xlsx")
        print("Archivo Excel guardado con exito")
        
    except Exception:
        print(f"Error al guardar  Archivo Excel")

def guardado_auto():
    reservas_json = {}
    for folio, datos in reservas.items():
        reservas_json[folio] = {
            "cliente": datos["cliente"],
            "sala": datos["sala"],
            "fecha": datos["fecha"].strftime(FORMATO_FECHA),
            "turno": datos["turno"],
            "evento": datos["evento"]
        }
    
    dict_json = {
        "Clientes": clientes,
        "Salas": salas,
        "Reservaciones": reservas_json
    }
    
    with open("datos_guardados.json", "w", encoding='utf-8') as archivo:
        json.dump(dict_json, archivo, indent=4, ensure_ascii= False)

def cargar_datos():
    global clientes, salas, reservas, salas_ocupadas
    try:
        with open("datos_guardados.json", "r") as archivo:
            lista_reconstruida = json.load(archivo)
            
            
            if not lista_reconstruida:
                clientes = {}  # clave: [apellidos, nombre]
                salas = {}     # clave: [nombre, cupo]
                reservas = {}  # folio: {cliente, sala, fecha, turno, evento}
                salas_ocupadas = {}
            else:
                while True:
                    print("Ya existe una versión anterior en el programa ¿desea recuperarla? (S/N) : ")

                    recuperar_input = input().strip()
                    
                    if not recuperar_input.isalpha():
                        print("Respuesta no válida")
                        continue
                    
                    if not (recuperar_input.upper() == "N" or recuperar_input.upper() == "S"):
                        print(f"La opción {recuperar_input} no está dentro de las opciones válidas. Intenta de nuevo.")
                        continue
                    
                    if recuperar_input.upper() == "N":
                        clientes = {}  # clave: [apellidos, nombre]
                        salas = {}     # clave: [nombre, cupo]
                        reservas = {}  # folio: {cliente, sala, fecha, turno, evento}
                        salas_ocupadas = {}
                        break
                    else:
                        clientes = {}
                        salas = {}
                        reservas = {}
                        salas_ocupadas = {}

                        clientes_data = lista_reconstruida["Clientes"]
                        for clave_str, datos_cliente in clientes_data.items():
                            clientes[int(clave_str)] = datos_cliente  # [apellidos, nombre]

                        salas_data = lista_reconstruida["Salas"]
                        for clave_str, datos_sala in salas_data.items():
                            salas[int(clave_str)] = datos_sala  # [nombre, cupo]

                        reservas_data = lista_reconstruida["Reservaciones"]
                        for folio_str, datos_reserva in reservas_data.items():
                            folio = int(folio_str)
                            
                            fecha_str = datos_reserva["fecha"]
                            fecha = datetime.datetime.strptime(fecha_str, FORMATO_FECHA).date()
                    
                            reservas[folio] = {
                        
                        "cliente": int(datos_reserva["cliente"]),
                        "sala": int(datos_reserva["sala"]),
                        "fecha": fecha,
                        "turno": datos_reserva["turno"],
                        "evento": datos_reserva["evento"]
                    }
                            sala = int(datos_reserva["sala"])
                            turno = datos_reserva["turno"]
                            salas_ocupadas[(sala, fecha, turno)] = folio
                
                        print("Estado anterior recuperado exitosamente.")
                        break
           
    except FileNotFoundError:
        clientes = {}  # clave: [apellidos, nombre]
        salas = {}     # clave: [nombre, cupo]
        reservas = {}  # folio: {cliente, sala, fecha, turno, evento}
        salas_ocupadas = {}
        print("No existe ninguna versión anterior")
    except json.JSONDecodeError:
        print("El archivo JSON está dañado o tiene formato incorrecto")
    except Exception:
        print(f"Ocurrió un error: {sys.exc_info()[0:2]}")

def menu_exportacion(fecha):
    while True:
            exportar = input("\nDesea exportar esta consulta a otro tipo de archivo? (S/N): ").strip()
            if not exportar.isalpha():
                print("Respuesta no valida")
                continue
            if not (exportar.upper() == "N" or  exportar.upper() == "S"):
                print(f"La opcion {exportar} no esta dentro de las opciones validas. Intenta de nuevo.")
                continue
            if exportar.upper() == "N":
                break
            else:
                print("\n--- MENÚ ---")
                print("[1] CSV")
                print("[2] JSON")
                print("[3] EXCEL")
                print("[4] CANCELAR")

                try:
                    op = input("\nElige una opción: ").strip()
                    if op == "1":   
                        print("\n")
                        export_csv(fecha)
                    elif op == "2":
                        print("\n")
                        export_json(fecha)
                    elif op == "3":
                        print("\n")
                        export_excel(fecha)
                    elif op == "4":
                        print("\n")
                        print("Cancelando Operacion...")
                        break
                    else:
                        print("Opción no válida.")
                except Exception:
                    print(f"Ocurrió un problema: {sys.exc_info()[0:2]}")
                    continue

#Opcion 1
def reservacion():
    try:
        if not mostrar_clientes():
            return
        while True:
            clave_cliente = input("\nDime tu clave de cliente (enter para salir): ").strip()
            if clave_cliente == "":
                print("Saliendo...\n\n")
                return
            if not clave_cliente.isdigit():
                print("\nClave inválida. Intenta de nuevo.\n\n")
                mostrar_clientes()
                continue
            elif int(clave_cliente) not in clientes:
                print("\nClave no registrada. Intenta de nuevo\n\n")
                mostrar_clientes()
                continue
            clave_cliente = int(clave_cliente)
            break

        while True:
            fecha_str = input("\nQue fecha le gustaria reservar (dd/mm/aaaa): ").strip()
            try:
                fecha = datetime.datetime.strptime(fecha_str, FORMATO_FECHA).date()
                if (fecha - fecha_actual).days < 2:
                    print("\nLa fecha debe ser al menos dos días después de hoy.\n")
                    continue
                break
            except ValueError:
                print("\nFormato de fecha incorrecto.\n\n")
                continue
            except Exception:
                print(f"\nOcurrió un problema: {sys.exc_info()[0:2]}\n\n")
                continue

        turnos = ["Mañana", "Tarde", "Noche"]
        print("\nTurnos disponibles: 1) Mañana 2) Tarde 3) Noche\n")
        while True:
            turno_op = input("\nSelecciona turno (1-3):  ").strip()
            if turno_op not in ["1", "2", "3"]:
                print("\nTurno inválido.\n")
                continue
            turno = turnos[int(turno_op)-1]
            break

        disponibles = mostrar_salas_disponibles(fecha, turno)
        if not disponibles:
            print("\nNo hay salas disponibles para ese turno y fecha.\n")
            return
        while True:
            clave_sala = input("\nClave de sala a reservar: ").strip()
            if not clave_sala.isdigit() or int(clave_sala) not in disponibles:
                print("\nClave de sala inválida.\n")
                continue
            clave_sala = int(clave_sala)
            break

        while True:
            evento = input("Nombre del evento: ").strip()
            if not evento:
                print("\nEl nombre del evento no puede estar vacío.\n")
                continue
            break

        folio = generar_folio()
        reservas[folio] = {
            "cliente": clave_cliente,
            "sala": clave_sala,
            "fecha": fecha,
            "turno": turno,
            "evento": evento
        }
        salas_ocupadas[(clave_sala, fecha, turno)] = folio
        print(f"\nReservación registrada con folio: {folio}\n")
    except ValueError:
        print("\nHubo un error al intentar ingresar un valor incorrecto\n")
        return
    except Exception:
        print(f"\nError en reservación: {sys.exc_info()[0:2]}\n ")

#Opcion 2
def editar_reservacion():
    try:
        print("Editar nombre de evento de una reservación.")
        fecha_inicial_str = input("Fecha inicial (dd/mm/aaaa): ").strip()
        fecha_final_str = input("Fecha final (dd/mm/aaaa): ").strip()
        try:
            fecha_inicial = datetime.datetime.strptime(fecha_inicial_str, FORMATO_FECHA).date()
            fecha_final = datetime.datetime.strptime(fecha_final_str, FORMATO_FECHA).date()
        except ValueError:
            print("Formato de fecha incorrecto.")
            return
        eventos = {folio: datos for folio, datos in reservas.items() if fecha_inicial <= datos["fecha"] <= fecha_final}
        if not eventos:
            print("No hay eventos en ese rango.")
            return
        print("FOLIO\tFECHA\t\tEVENTO")
        print("-" * 40)
        for folio, datos in eventos.items():
            print(f"{folio}\t{datos['fecha'].strftime(FORMATO_FECHA)}\t{datos['evento']}")
        while True:
            folio_seleccionado = input("Folio a editar (enter para cancelar): ").strip()
            if folio_seleccionado == "":
                print("Cancelado.")
                return
            if not folio_seleccionado.isdigit() or int(folio_seleccionado) not in eventos:
                print("Folio inválido.")
                continue
            folio_seleccionado = int(folio_seleccionado)
            break
        while True:
            nuevo_nombre = input("Nuevo nombre del evento: ").strip()
            if not nuevo_nombre:
                print("El nombre del evento no puede estar vacío.")
                continue
            reservas[folio_seleccionado]["evento"] = nuevo_nombre
            print("Evento actualizado.")
            break
    except Exception:
        print(f"Error al editar reservación: {sys.exc_info()[0:2]} ")

#Opcion 3
def consultar_reservaciones():
    try:
        fecha_str = input("Fecha a consultar (dd/mm/aaaa): ").strip()
        fecha = datetime.datetime.strptime(fecha_str, FORMATO_FECHA).date()
        
        print("\n\nFOLIO\tSALA\tTURNO\tEVENTO\t\tCLIENTE")
        print("-" * 60)
        for folio, datos in reservas.items():
            if datos["fecha"] == fecha:
                sala = salas[datos["sala"]][0]
                cliente = ([datos["cliente"]][::-1])
                print(f"{folio:<8}{sala:<8}{datos['turno']:<8}{datos['evento']:<16}{cliente}")
    except ValueError:
            print("Formato de fecha incorrecto.")
            return    
    except Exception:
        print(f"Error al consultar reservaciones: {sys.exc_info()[0:2]} ")
    else:
        menu_exportacion(fecha)
    

#Opcion 4
def new_cliente():
    try:
        clave = max(clientes.keys(), default=100) + 1
        nombre = input("Nombre: ").strip()
        if not (nombre.isalpha() ):
            print("\nEl nombre debe contener solo letras.\n")
            return
        apellidos = input("Apellidos: ").strip()
        if not (apellidos.isalpha()):
            print("\nLos apellidos deben contener solo letras.\n")
            return
        clientes[clave] = [apellidos, nombre]
    except Exception:
        print(f"\nError al registrar cliente: {sys.exc_info()[0:2]}\n")
    else:
        print(f"\nCliente registrado con clave: {clave}\n\n")

#Opcion 5
def new_sala():
    try:
        clave = max(salas.keys(), default=0) + 1
        nombre = input("Nombre de la sala: ").strip()
        if not nombre.isalpha():
            print("\nEl nombre de la sala debe contener solo letras.\n")
            return
        cupo = input("Capacidad de la sala: ").strip()
        if not cupo.isdigit() or int(cupo) <= 0:
            print("\nEl cupo debe ser un número entero positivo.\n")
            return 
        salas[clave] = [nombre, int(cupo)]
    except ValueError:
        print("\nEl cupo debe ser un numero entero positivo\n")
        return
    except Exception:
        print(f"\nError al registrar sala: {sys.exc_info()[0:2]}\n")
    else:
        print(f"\nSala registrada con clave: {clave}\n\n")

def main():
    cargar_datos()
    while True:
        print("\n--- MENÚ ---")
        print("[1] Registrar reservación")
        print("[2] Editar reservación")
        print("[3] Consultar reservaciones")
        print("[4] Registrar cliente")
        print("[5] Registrar sala")
        print("[6] Salir")
        try:
            op = input("\nElige una opción: ").strip()
            if op == "1":
                print("\n")
                reservacion()
            elif op == "2":
                print("\n")
                editar_reservacion()
            elif op == "3":
                print("\n")
                consultar_reservaciones()
            elif op == "4":
                print("\n")
                new_cliente()
            elif op == "5":
                print("\n")
                new_sala()
            elif op == "6":
                print("\n")
                try:
                    confirmacion = input("Esta seguro que desea salir? (S/N): ").strip()
                    if not confirmacion.isalpha():
                        continue
                    if not (confirmacion.upper() == "N" or  confirmacion.upper() == "S"):
                        print(f"La opcion {confirmacion} no esta dentro de las opciones validas. Intenta de nuevo.")
                        continue
                    if confirmacion.upper()== "N":
                        continue
                    else:
                        print("Saliendo del programa...")
                        guardado_auto()
                        break
                except Exception:
                    print(f"Ocurrió un problema: {sys.exc_info()[0:2]}")
                    continue
            else:
                print("Opción no válida.")

        except ValueError:
            print("La opcion proporcionada no es valida, intente de nuevo...")
            continue
        except Exception:
            print(f"Ocurrió un problema: {sys.exc_info()[0:2]}")
            continue


if __name__ == "__main__":
    main()
